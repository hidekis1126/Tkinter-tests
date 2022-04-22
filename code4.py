import tkinter
from pathlib import Path
from tkinter import filedialog
import openpyxl

class SaveApplication(tkinter.Frame):
    def __init__(self, root=None):
        super().__init__(root, width=420, height=720,
                         borderwidth=4, relief='groove')
        self.root = root
        self.pack()
        self.pack_propagate(0)
        self.create_widgets()
        self.file_name = Path.cwd() / 'app_data.xlsx'

    def create_widgets(self):

        self.cell_positions = []
        self.cell_texts = []

        for i in range(10):
            # メッセージ出力
            cell_label = tkinter.Message(self)
            cell_label['text'] = 'セル'
            cell_label.place(x='1', y=str(i * 22))

            # テキストボックス
            position_box = tkinter.Entry(self)
            position_box['width'] = 4
            position_box.place(x='40', y=str(i * 22))

            # メッセージ出力
            text_label = tkinter.Message(self)
            text_label['text'] = '値 '
            text_label.place(x='100', y=str(i * 22))

            # テキストボックス
            text_box = tkinter.Entry(self)
            text_box['width'] = 20
            text_box.place(x='150', y=str(i * 22))

            self.cell_positions.append(position_box)
            self.cell_texts.append(text_box)

        # 閉じるボタン
        quit_btn = tkinter.Button(self)
        quit_btn['text'] = "閉じる"
        quit_btn['command'] = self.root.destroy
        quit_btn.pack(side="bottom")

        # 完了メッセージ出力
        self.message = tkinter.Message(self)
        self.message.place(x='180', y=310)

        # 注意書きメッセージ出力
        note_message = tkinter.Message(self)
        note_message['width'] = 300
        note_message['text'] = '※セルはA1:G10の範囲で設定してください'
        note_message.place(x='20', y=225)

        # 保存ボタン
        submit_btn = tkinter.Button(self)
        submit_btn['text'] = "保存"
        submit_btn['command'] = self.save_data
        submit_btn.place(x='180', y=250)

        # データ読み込み
        submit_btn = tkinter.Button(self)
        submit_btn['text'] = "読み込み"
        submit_btn['command'] = self.load_data
        submit_btn.place(x='170', y=280)

        # 読み取りデータ出力
        self.data = tkinter.Message(self)
        self.data['width'] = 300
        self.data.pack(side='bottom', expand=1)


    def save_data(self):
        wb = openpyxl.load_workbook(self.file_name)
        ws = wb.worksheets[0]

        for i in range(10):
            text = self.cell_texts[i].get()
            if position := self.cell_positions[i].get():
                ws[position].value = text

        wb.save(self.file_name)
        self.message['text'] = '保存完了'

    def load_data(self):
        wb = openpyxl.load_workbook(self.file_name)
        ws = wb.worksheets[0]

        values = []
        for row in range(1, 10):
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
                if value := ws[f'{col}{row}'].value:
                    values.append(value)
        if values:
            self.data['text'] = '  '.join(values)


root = tkinter.Tk()
root.title('hideki4 アプリ')
root.geometry('450x750')
save_app = SaveApplication(root=root)
save_app.mainloop()
